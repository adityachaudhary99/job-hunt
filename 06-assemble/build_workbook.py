"""Assemble the final job-hunt.xlsx workbook and tracker.md.

Sheets (in order):
  1. Catalog              — all 90 original tabs (read-only reference)
  2. Sources              — filtered catalog: career pages + aggregators + referral forms
  3. Companies            — full deduped company list (yc + a16z + tabs + wordlist) with discovered ATS
  4. Shortlist            — top-ranked matched postings (role + geo + recency + hiring boost)
  5. Postings (raw)       — every scraped posting
  6. Postings (matched)   — postings whose title/department matched one of the role keywords (sorted newest-first)
  7. Tracker              — empty application log (you fill rows as you apply)

Also writes tracker.md with a compact version of the data.
"""
from __future__ import annotations

import csv
import datetime as _dt
import re
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from paths import TABS_CATALOG, POSTINGS, XLSX, TRACKER_MD, COMPANIES_WITH_ATS

CATALOG_CSV = TABS_CATALOG
POSTINGS_CSV = POSTINGS
COMPANIES_CSV = COMPANIES_WITH_ATS
XLSX_PATH = XLSX
MD_PATH = TRACKER_MD

SOURCE_TYPES = {
    "company-career-page",
    "job-board-aggregator",
    "referral-or-form",
    "specific-job-posting",
    "company-landing-page",
}

TRACKER_COLUMNS = [
    "date_added", "company", "role_title", "apply_url", "source_url",
    "ats", "location", "geo_tags", "role_tags", "match_score",
    "status", "date_applied", "response_date", "resume_version",
    "cover_letter", "referral", "notes",
]

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
MD_POSTINGS_CAP = 300  # cap markdown matched-postings table to keep file scannable


def load_csv(path: Path) -> list[dict]:
    if not path.exists():
        return []
    with path.open(encoding="utf-8") as f:
        return list(csv.DictReader(f))


def parse_iso_date(s: str) -> _dt.datetime | None:
    """Best-effort parse. Returns naive UTC datetime (tzinfo stripped) so all
    comparisons against `datetime.utcnow()` are consistent regardless of which
    ATS produced the string."""
    if not s:
        return None
    s = s.strip()
    # Lever uses millisecond Unix timestamps
    if s.isdigit() and len(s) >= 10:
        try:
            ts = int(s)
            if ts > 10**12:
                ts //= 1000
            return _dt.datetime.utcfromtimestamp(ts)
        except Exception:
            return None
    for fmt in ("%Y-%m-%dT%H:%M:%S.%fZ", "%Y-%m-%dT%H:%M:%S%z",
                "%Y-%m-%dT%H:%M:%SZ", "%Y-%m-%dT%H:%M:%S",
                "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            dt = _dt.datetime.strptime(s.replace("+00:00", "Z"), fmt)
        except ValueError:
            continue
        # Strip tzinfo so downstream arithmetic stays naive-vs-naive
        if dt.tzinfo is not None:
            dt = dt.astimezone(_dt.timezone.utc).replace(tzinfo=None)
        return dt
    return None


def dedupe_postings(postings: list[dict]) -> list[dict]:
    """Same role can appear on multiple boards. Group by (company-lower, title-lower, location-lower);
    keep the entry with the newest published_at."""
    bucket: dict[tuple[str, str, str], dict] = {}
    for p in postings:
        key = (
            (p.get("source_company") or "").lower().strip(),
            re.sub(r"\s+", " ", (p.get("posting_title") or "")).lower().strip(),
            (p.get("location") or "").lower().strip(),
        )
        if not key[1]:
            continue
        existing = bucket.get(key)
        if existing is None:
            bucket[key] = p
            continue
        # Keep whichever has the newer date
        new_dt = parse_iso_date(p.get("published_at", ""))
        old_dt = parse_iso_date(existing.get("published_at", ""))
        if new_dt and (not old_dt or new_dt > old_dt):
            bucket[key] = p
    return list(bucket.values())


# Preferences pulled from memory file user_job_search.md (2026-05-17 update).
PREFERRED_GEOS = {"India", "Remote", "US", "Europe", "Japan"}
PREFERRED_ROLES = {"data", "ai-ml", "backend", "fullstack", "software", "fde", "early-career", "infra"}
PRIMARY_ROLES = {"fde", "data", "ai-ml", "infra"}  # FDE + AI Infra + Data Eng

# Which ATSes are allowed into the Shortlist sheet (raw/matched sheets always
# contain everything regardless). Edit this set to drop ATSes whose apply UX
# you dislike or whose postings you don't want to wade through. Set to None or
# empty set for no filter.
# Default: all 5 we scrape. Common edits:
#   {"Ashby", "Greenhouse", "Lever"}  — only the top-3 with cleanest apply flows
#   {"Ashby", "Greenhouse"}           — most frictionless apply experience
SHORTLIST_ATSES: set[str] | None = {"Ashby", "Greenhouse", "Lever", "SmartRecruiters", "Recruitee"}

# Seniority inference from title — user has ~1 YoE, so titles imply YoE bands.
# Order matters: first match wins, so most-specific patterns come first.
SENIORITY_PATTERNS = [
    ("intern",       re.compile(r"\b(intern|internship)\b", re.I)),
    ("new-grad",     re.compile(r"\b(new grad|new-grad|graduate program|university grad|fresh(er)?)\b", re.I)),
    ("junior",       re.compile(r"\b(junior|jr\.?|entry[- ]level|associate engineer|trainee|early career|early[- ]career)\b", re.I)),
    ("distinguished",re.compile(r"\b(distinguished|fellow)\b", re.I)),
    ("vp",           re.compile(r"\b(vp|vice president|chief|cto|cio|cpo|cso|head of)\b", re.I)),
    ("director",     re.compile(r"\b(director)\b", re.I)),
    ("manager",      re.compile(r"\b(manager|people manager|engineering manager|em\b)\b", re.I)),
    ("principal",    re.compile(r"\b(principal|architect)\b", re.I)),
    ("staff",        re.compile(r"\b(staff)\b", re.I)),
    ("senior",       re.compile(r"\b(senior|sr\.?|lead engineer|tech lead)\b", re.I)),
    ("mid",          re.compile(r"\b(mid[- ]level|mid level|ii|iii)\b", re.I)),
]

# Scoring penalty by inferred seniority (negative = pushes posting down)
SENIORITY_SCORE = {
    # User has ~1 YoE *post-grad*, currently employed FDE. So:
    #   - junior / new-grad = sweet spot
    #   - mid = often reachable
    #   - untagged ("Software Engineer" etc.) = neutral baseline
    #   - intern = neutral (most are student-only, but a few open to early-career; let other signals decide)
    #   - senior+ = pushed down
    "new-grad":      +1.5,
    "junior":        +1.5,
    "mid":           +0.5,
    "":              0.0,
    "intern":        0.0,
    "senior":        -2.0,
    "manager":       -3.0,
    "staff":         -4.0,
    "principal":     -5.0,
    "distinguished": -5.0,
    "director":      -5.0,
    "vp":            -6.0,
}


PHD_PATTERN = re.compile(r"\b(ph\.?d|doctorate|doctoral)\b", re.I)


def infer_seniority(title: str) -> str:
    if not title:
        return ""
    for label, pat in SENIORITY_PATTERNS:
        if pat.search(title):
            return label
    return ""


def title_phd_required(title: str) -> bool:
    return bool(title and PHD_PATTERN.search(title))


def score_posting(p: dict, companies_by_name: dict[str, dict]) -> float:
    """Heuristic ~-3..+11 score for a matched posting.

    Components:
      role_match     +0..+3    number of preferred role tags hit (capped)
      primary_bonus  +0..+2    bonus for hitting PRIMARY_ROLES (fde, data, ai-ml, infra)
      geo_match      +0..+2    number of preferred geo tags hit (capped)
      recency        +0..+3    linear from 0 (>180d) to 3 (today)
      yc_hiring      +0..+1    1 if company is YC + is_hiring=true
      seniority      -6..+1.5  inferred from title; user has ~1 YoE so senior+ titles get penalised
    """
    role_tags = set((p.get("role_tags") or "").split(";")) - {""}
    geo_tags = set((p.get("geo_tags") or "").split(";")) - {""}
    role_score = min(len(role_tags & PREFERRED_ROLES), 3)
    primary_bonus = min(len(role_tags & PRIMARY_ROLES), 2)
    geo_score = min(len(geo_tags & PREFERRED_GEOS), 2)

    pub_dt = parse_iso_date(p.get("published_at", ""))
    if pub_dt:
        age_days = (_dt.datetime.utcnow() - pub_dt).days
        recency = max(0.0, min(3.0, 3 * (1 - age_days / 180)))
    else:
        recency = 0.0

    yc_boost = 0
    co = companies_by_name.get((p.get("source_company") or "").lower())
    if co and co.get("is_hiring") == "true" and "yc" in (co.get("source") or ""):
        yc_boost = 1

    seniority = SENIORITY_SCORE.get(p.get("_seniority", ""), 0.0)
    phd_penalty = -3.0 if title_phd_required(p.get("posting_title", "")) else 0.0

    return round(role_score + primary_bonus + geo_score + recency + yc_boost + seniority + phd_penalty, 2)


def write_sheet(ws, rows: list[dict], headers: list[str]) -> None:
    ws.append(headers)
    for r in rows:
        ws.append([r.get(h, "") for h in headers])
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    for col_i, header in enumerate(headers, start=1):
        max_len = len(header)
        for r in rows[:500]:  # sample for width calc only — speeds up large sheets
            v = str(r.get(header, "") or "")
            if len(v) > max_len:
                max_len = len(v)
        ws.column_dimensions[get_column_letter(col_i)].width = min(max(max_len + 2, 10), 60)


def build_xlsx(catalog: list[dict], postings: list[dict], companies: list[dict]) -> None:
    wb = Workbook()

    # Sheet 1: Catalog
    ws = wb.active
    ws.title = "Catalog"
    catalog_headers = ["index", "company", "tab_type", "ats_platform", "geo_hint", "title", "url", "domain", "notes"]
    write_sheet(ws, catalog, catalog_headers)

    # Sheet 2: Sources
    sources = [r for r in catalog if r.get("tab_type") in SOURCE_TYPES]
    write_sheet(wb.create_sheet("Sources"), sources, catalog_headers)

    # Sheet 3: Companies (full master list, sorted: ATS-discovered first, then alpha)
    companies_sorted = sorted(companies, key=lambda c: (not bool(c.get("ats")),
                                                       c.get("company_name", "").lower()))
    company_headers = ["company_name", "ats", "ats_slug", "is_hiring", "source",
                       "industry", "location", "batch", "website", "source_url", "slug_candidates"]
    write_sheet(wb.create_sheet("Companies"), companies_sorted, company_headers)

    # Dedupe postings across boards before producing the curated sheets.
    # Postings are already tagged with seniority by main(); dedupe preserves dict identity.
    deduped = dedupe_postings(postings)
    print(f"  deduped {len(postings)} -> {len(deduped)} postings")

    # Sort newest-first using parsed published_at
    def sort_key(p: dict):
        dt = parse_iso_date(p.get("published_at", ""))
        return (dt is None, -(dt.timestamp() if dt else 0))

    deduped_sorted = sorted(deduped, key=sort_key)

    # Sheet 4: Shortlist — top postings by score, optionally ATS-filtered
    companies_by_name = {(c.get("company_name") or "").lower(): c for c in companies}
    matched = [p for p in deduped_sorted if (p.get("role_tags") or "").strip()]
    scored = [(score_posting(p, companies_by_name), p) for p in matched]
    scored.sort(key=lambda sp: -sp[0])

    # ATS allowlist filter — applied only to Shortlist, not to matched/raw sheets
    if SHORTLIST_ATSES:
        before_atses = len(scored)
        scored_for_shortlist = [(s, p) for s, p in scored if p.get("ats") in SHORTLIST_ATSES]
        dropped = before_atses - len(scored_for_shortlist)
        if dropped:
            print(f"  ATS filter dropped {dropped} matched postings from shortlist "
                  f"(allowlist: {sorted(SHORTLIST_ATSES)})")
    else:
        scored_for_shortlist = scored

    shortlist_headers = ["score", "seniority"] + ["source_company", "ats", "posting_title", "department",
                                       "location", "role_tags", "geo_tags", "published_at",
                                       "apply_url", "source_url"]
    SHORTLIST_TOP_N = 200
    shortlist_rows = []
    for score, p in scored_for_shortlist[:SHORTLIST_TOP_N]:
        row = {h: p.get(h, "") for h in shortlist_headers if h != "score"}
        row["score"] = score
        shortlist_rows.append(row)
    write_sheet(wb.create_sheet("Shortlist"), shortlist_rows, shortlist_headers)
    print(f"  built shortlist with {len(shortlist_rows)} top postings")

    # Sheet 5: Postings (raw) — full unfiltered list, newest first
    posting_headers = ["source_company", "ats", "seniority", "posting_title", "department", "location",
                       "role_tags", "geo_tags", "employment_type", "published_at",
                       "apply_url", "source_url", "posting_id"]
    write_sheet(wb.create_sheet("Postings (raw)"), sorted(postings, key=sort_key), posting_headers)

    # Sheet 6: Postings (matched) — deduped, sorted newest first
    write_sheet(wb.create_sheet("Postings (matched)"), matched, posting_headers)

    # Sheet 6: Tracker template
    ws_tracker = wb.create_sheet("Tracker")
    write_sheet(ws_tracker, [], TRACKER_COLUMNS)
    sample = {
        "date_added": "2026-05-15",
        "company": "Example Co",
        "role_title": "Data Engineer",
        "apply_url": "https://example.com/jobs/123",
        "source_url": "https://example.com/careers",
        "ats": "Greenhouse",
        "location": "Bangalore",
        "geo_tags": "India",
        "role_tags": "data",
        "match_score": "3",
        "status": "interested",
        "date_applied": "",
        "response_date": "",
        "resume_version": "v3-data-eng",
        "cover_letter": "no",
        "referral": "no",
        "notes": "Sample row — delete after first real entry.",
    }
    ws_tracker.append([sample[h] for h in TRACKER_COLUMNS])
    for col in range(1, len(TRACKER_COLUMNS) + 1):
        ws_tracker.cell(row=2, column=col).font = Font(italic=True, color="888888")

    wb.save(XLSX_PATH)
    print(f"Wrote {XLSX_PATH}")


def build_markdown(catalog: list[dict], postings: list[dict], companies: list[dict]) -> None:
    sources = [r for r in catalog if r.get("tab_type") in SOURCE_TYPES]
    deduped = dedupe_postings(postings)

    def sort_key(p: dict):
        dt = parse_iso_date(p.get("published_at", ""))
        return (dt is None, -(dt.timestamp() if dt else 0))

    companies_by_name = {(c.get("company_name") or "").lower(): c for c in companies}
    matched = [p for p in deduped if (p.get("role_tags") or "").strip()]
    scored = sorted(((score_posting(p, companies_by_name), p) for p in matched),
                    key=lambda sp: -sp[0])
    if SHORTLIST_ATSES:
        scored = [(s, p) for s, p in scored if p.get("ats") in SHORTLIST_ATSES]
    matched_by_score = [p for _, p in scored]
    ats_companies = [c for c in companies if c.get("ats")]

    lines: list[str] = []
    lines.append("# Job Hunt Tracker")
    lines.append("")
    lines.append("_Auto-generated. Re-run scripts in 01-ingest → 06-assemble to refresh._")
    lines.append("")
    lines.append(f"- Tabs catalogued: **{len(catalog)}**")
    lines.append(f"- Companies in master list (yc + a16z + tabs + wordlist): **{len(companies)}**")
    lines.append(f"- Companies with discovered ATS: **{len(ats_companies)}**")
    lines.append(f"- Postings scraped (raw): **{len(postings)}**")
    lines.append(f"- Postings after dedup: **{len(deduped)}**")
    lines.append(f"- Postings matching role keywords: **{len(matched)}**")
    lines.append("")

    # ATS-by-source summary
    lines.append("## Companies by ATS")
    lines.append("")
    from collections import Counter
    by_ats = Counter(c.get("ats", "") for c in companies if c.get("ats"))
    lines.append("| ATS | Companies discovered |")
    lines.append("|---|---|")
    for k, v in by_ats.most_common():
        lines.append(f"| {k} | {v} |")
    lines.append("")

    # Sources table — grouped by type
    lines.append("## Sources (from your original tabs)")
    lines.append("")
    from collections import defaultdict
    by_type = defaultdict(list)
    for s in sources:
        by_type[s.get("tab_type", "")].append(s)
    for tab_type in sorted(by_type):
        lines.append(f"### {tab_type}  _({len(by_type[tab_type])})_")
        lines.append("")
        lines.append("| Company | Geo hint | ATS | URL |")
        lines.append("|---|---|---|---|")
        for s in by_type[tab_type]:
            url = s.get("url", "")
            short = url if len(url) <= 80 else url[:77] + "..."
            lines.append(f"| {s.get('company','')} | {s.get('geo_hint','')} | {s.get('ats_platform','')} | [{short}]({url}) |")
        lines.append("")

    # Shortlist (top by score)
    lines.append(f"## Shortlist — top {min(MD_POSTINGS_CAP, len(matched_by_score))} matched postings by score")
    lines.append("")
    lines.append("Ranked by role match + geo match + recency + YC-hiring boost. Full ranked list in `job-hunt.xlsx` → _Shortlist_ sheet.")
    lines.append("")
    lines.append("| Score | Company | Role tag | Title | Location | Geo tag | Apply |")
    lines.append("|---|---|---|---|---|---|---|")
    for p in matched_by_score[:MD_POSTINGS_CAP]:
        s = score_posting(p, companies_by_name)
        lines.append(
            f"| {s} "
            f"| {p.get('source_company','')} "
            f"| {p.get('role_tags','')} "
            f"| {p.get('posting_title','').replace('|','\\|')} "
            f"| {p.get('location','').replace('|','\\|')} "
            f"| {p.get('geo_tags','')} "
            f"| [link]({p.get('apply_url','')}) |"
        )
    lines.append("")

    # Tracker template
    lines.append("## Application Tracker")
    lines.append("")
    lines.append("Fill a row each time you apply. Status values: `interested`, `applied`, `phone`, `onsite`, `offer`, `rejected`, `withdrew`.")
    lines.append("")
    lines.append("| " + " | ".join(TRACKER_COLUMNS) + " |")
    lines.append("|" + "|".join("---" for _ in TRACKER_COLUMNS) + "|")
    lines.append("| " + " | ".join("" for _ in TRACKER_COLUMNS) + " |")
    lines.append("")

    MD_PATH.write_text("\n".join(lines), encoding="utf-8")
    print(f"Wrote {MD_PATH}")


def main() -> None:
    catalog = load_csv(CATALOG_CSV)
    postings = load_csv(POSTINGS_CSV)
    companies = load_csv(COMPANIES_CSV)
    print(f"Loaded {len(catalog)} tabs, {len(postings)} postings, {len(companies)} companies")

    # Tag every posting with inferred seniority once, here, so xlsx + markdown agree.
    from collections import Counter
    sen = Counter()
    for p in postings:
        s = infer_seniority(p.get("posting_title", ""))
        p["seniority"] = s
        p["_seniority"] = s
        sen[s] += 1
    print(f"  seniority tags: {dict(sen)}")

    build_xlsx(catalog, postings, companies)
    build_markdown(catalog, postings, companies)


if __name__ == "__main__":
    main()
