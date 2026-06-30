"""Export job-hunt's top-scored matched postings as a Markdown task list.

Writes `- [ ] <url> | <company> | <role> | score=X.X` lines into a downstream
`data/pipeline.md` inbox under `## Pending`, ready to feed a per-job evaluator
or agent. Idempotent: URLs already present (in `## Pending` or `## Processed`)
are skipped, so re-runs are safe.

Two source modes:

  --source xlsx  (default)
      Read the `Shortlist` sheet in `06-assemble/job-hunt.xlsx` (rows already
      carry `score`). Cheap, fast, deterministic.

  --source csv
      Read raw `05-scrape/postings.csv` + `04-discover/companies_with_ats.csv`
      and score in-process via `score_posting` from `06-assemble/build_workbook.py`.
      Use when you want fresh scoring without rebuilding the workbook.

Target dir: `$HANDOFF_DIR`, else `--out`, else a local `./bridge-out`.

USAGE
=====

    python 07-bridge/bridge_to_evaluator.py --limit 30
    python 07-bridge/bridge_to_evaluator.py --source csv --min-score 6.0
    HANDOFF_DIR=/path/to/evaluator python 07-bridge/bridge_to_evaluator.py
    python 07-bridge/bridge_to_evaluator.py --out ./bridge-out
"""
from __future__ import annotations

import argparse
import csv
import os
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from paths import POSTINGS, COMPANIES_WITH_ATS, XLSX  # noqa: E402

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "06-assemble"))
from build_workbook import (  # noqa: E402
    dedupe_postings,
    score_posting,
)


def default_handoff_dir() -> Path:
    """Where to write the hand-off file. Override with $HANDOFF_DIR or --out.

    Defaults to a local ``bridge-out/`` so a fresh clone works out of the box;
    point it at your own downstream evaluator's dir when you have one.
    """
    env = os.environ.get("HANDOFF_DIR")
    if env:
        return Path(env)
    return Path(__file__).resolve().parent.parent / "bridge-out"


PENDING_HEADER = "## Pending"
PROCESSED_HEADER = "## Processed"

SHORTLIST_SHEET = "Shortlist"


# ---------------------------------------------------------------- source: xlsx


def read_from_xlsx(xlsx_path: Path, top: int) -> list[tuple[float, dict]]:
    """Read top-N rows from the Shortlist sheet (already scored, already sorted)."""
    from openpyxl import load_workbook  # local import to keep csv-only runs lighter

    if not xlsx_path.exists():
        raise FileNotFoundError(
            f"workbook not found: {xlsx_path}\n"
            f"  → run 06-assemble/build_workbook.py first, "
            f"or use --source csv"
        )
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    if SHORTLIST_SHEET not in wb.sheetnames:
        wb.close()
        raise KeyError(
            f"sheet '{SHORTLIST_SHEET}' missing from {xlsx_path.name}.\n"
            f"  → rebuild the workbook (06-assemble/build_workbook.py) "
            f"or use --source csv"
        )
    ws = wb[SHORTLIST_SHEET]
    rows_iter = ws.iter_rows(values_only=True)
    headers = [h or "" for h in next(rows_iter)]

    scored: list[tuple[float, dict]] = []
    for raw in rows_iter:
        rec = dict(zip(headers, raw))
        score = float(rec.get("score") or 0)
        scored.append((score, rec))
        if len(scored) >= top:
            break
    wb.close()
    print(f"[xlsx] read {len(scored)} rows from '{SHORTLIST_SHEET}' sheet")
    return scored


# ---------------------------------------------------------------- source: csv


def load_csv(path: Path) -> list[dict]:
    with path.open(encoding="utf-8") as f:
        return list(csv.DictReader(f))


def is_matched(p: dict) -> bool:
    """A posting is 'matched' if it has at least one role_tag."""
    return bool((p.get("role_tags") or "").strip())


def read_from_csv(top: int) -> list[tuple[float, dict]]:
    """Read postings.csv + companies_with_ats.csv, dedupe, score, return top-N."""
    postings = load_csv(POSTINGS)
    companies = load_csv(COMPANIES_WITH_ATS)
    companies_by_name = {(c.get("company_name") or "").lower(): c for c in companies}
    print(f"[csv] loaded {len(postings)} postings, {len(companies)} companies")

    matched = [p for p in postings if is_matched(p)]
    matched = dedupe_postings(matched)
    print(f"[csv] matched + deduped: {len(matched)}")

    scored = sorted(
        ((score_posting(p, companies_by_name), p) for p in matched),
        key=lambda x: x[0],
        reverse=True,
    )
    top_rows = scored[:top]
    if top_rows:
        print(f"[csv] top-{top} score range: {top_rows[-1][0]:.2f} – {top_rows[0][0]:.2f}")
    return top_rows


# ---------------------------------------------------------------- write side


def parse_existing(pipeline_path: Path) -> tuple[list[str], list[str], set[str]]:
    """Return (pending_lines, processed_lines, all_urls_lowercased)."""
    if not pipeline_path.exists():
        return [], [], set()
    text = pipeline_path.read_text(encoding="utf-8")
    pending: list[str] = []
    processed: list[str] = []
    section = None
    for line in text.splitlines():
        stripped = line.strip()
        if stripped.startswith("## "):
            if stripped == PENDING_HEADER:
                section = "pending"
            elif stripped == PROCESSED_HEADER:
                section = "processed"
            else:
                section = None
            continue
        if section == "pending" and stripped:
            pending.append(line)
        elif section == "processed" and stripped:
            processed.append(line)

    urls: set[str] = set()
    for line in pending + processed:
        for tok in line.split():
            if tok.startswith("http"):
                urls.add(tok.lower().rstrip("|").strip(",.;"))
                break
    return pending, processed, urls


def build_line(p: dict, score: float) -> str:
    url = (p.get("apply_url") or "").strip()
    company = (p.get("source_company") or "").strip()
    role = (p.get("posting_title") or "").strip()
    bits = [url]
    if company:
        bits.append(company)
    if role:
        bits.append(role)
    bits.append(f"score={score:.1f}")
    return f"- [ ] {' | '.join(bits)}"


# ---------------------------------------------------------------- main


def main() -> int:
    ap = argparse.ArgumentParser(
        formatter_class=argparse.RawDescriptionHelpFormatter,
        description=__doc__,
    )
    ap.add_argument(
        "--source", choices=("xlsx", "csv"), default="xlsx",
        help="where to read scored postings from (default: xlsx -> Shortlist sheet)",
    )
    ap.add_argument(
        "--top", type=int, default=200,
        help="rank all matched postings by score; consider top-N (default 200)",
    )
    ap.add_argument(
        "--limit", type=int, default=None,
        help="cap NEW URLs appended in this run (default: unlimited)",
    )
    ap.add_argument(
        "--min-score", type=float, default=None,
        help="only bridge rows with score >= this",
    )
    ap.add_argument(
        "--out", default=None,
        help="target dir for the hand-off (default: $HANDOFF_DIR, else ./bridge-out)",
    )
    args = ap.parse_args()

    handoff_dir = Path(args.out) if args.out else default_handoff_dir()
    pipeline_md = handoff_dir / "data" / "pipeline.md"

    if args.source == "xlsx":
        scored = read_from_xlsx(XLSX, args.top)
    else:
        scored = read_from_csv(args.top)

    if args.min_score is not None:
        before = len(scored)
        scored = [(s, p) for (s, p) in scored if s >= args.min_score]
        print(f"after min-score filter (>= {args.min_score}): {len(scored)} (was {before})")

    pipeline_md.parent.mkdir(parents=True, exist_ok=True)
    pending, processed, seen = parse_existing(pipeline_md)

    new_lines: list[str] = []
    skipped_dupes = 0
    skipped_no_url = 0
    for score, p in scored:
        url = (p.get("apply_url") or "").strip()
        if not url:
            skipped_no_url += 1
            continue
        if url.lower() in seen:
            skipped_dupes += 1
            continue
        new_lines.append(build_line(p, score))
        seen.add(url.lower())
        if args.limit is not None and len(new_lines) >= args.limit:
            break

    print(
        f"new URLs to append: {len(new_lines)}   "
        f"(skipped {skipped_dupes} duplicates, {skipped_no_url} without apply_url)"
    )

    pending_all = pending + new_lines

    parts: list[str] = [PENDING_HEADER, ""]
    parts.extend(pending_all if pending_all else ["<!-- empty -->"])
    parts.extend(["", PROCESSED_HEADER, ""])
    parts.extend(processed if processed else ["<!-- empty -->"])
    parts.append("")
    pipeline_md.write_text("\n".join(parts), encoding="utf-8")
    print(f"wrote {pipeline_md}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
